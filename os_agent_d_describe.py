# os_agent_d_describe.py
#
# 每阶段 Cursor 式管线（固定）:
#   ① Plan     — 阶段提示词 + 启发式 plan + LLM 自主探索 → 锁定计划与 execution_steps（须按序执行）
#   ② Execute  — System + 单条 Human（计划+任务合并）后 ReAct 写章节
#   ③ Review（可选）— DESCRIBE_STAGE_REVIEW=1 且本阶段为 JSON-QA 并成功校验时：仅将
#      「describe_stage_qa 题单」+「coerce_answers_payload_by_stage_qa 覆写前的 JSON」
#      送审；不含工具摘录。01_overview / 10_history 不审。
#      侧车：_per_stage/<stage_id>_review.json（失败见 *_review_error.json）。
#
# API Key / MODEL_NAME 等仍从仓库根目录 .env 读取（load_dotenv）。
#
import copy
import json
import os
import re
import sys
import time
import logging
from datetime import datetime

from dotenv import load_dotenv

# 尽早加载：override=True 使 .env 覆盖系统/用户环境变量里误设的 HF_ENDPOINT 等（默认 load_dotenv 不覆盖）
load_dotenv(override=True)

import langchain
from langchain_core.messages import AIMessage, HumanMessage, ToolMessage, SystemMessage

from core.agent_builder import (
    build_chat_model,
    build_executor_agent,
    build_planner_agent,
    SYSTEM_PROMPT,
    DESCRIBE_SYSTEM_PROMPT_JSON,
)
from core.utils import repo_name_from_url, format_tool_call_summary, format_tool_result_summary
from core.error_handling import ErrorType, RetryConfig, classify_error, calculate_backoff, ErrorTracker
from core.per_types import StageState
from core.per_planner import (
    apply_llm_plan_overlay,
    build_repo_profile,
    build_dynamic_context,
    ensure_execution_steps,
    plan_stage,
    render_plan_context,
)
from core.per_executor import extract_stage_artifacts
from core.describe_stage_qa import load_stage_qa, list_question_ids
from core.describe_stage_review import (
    build_stage_qa_question_sheet,
    describe_stage_review_applies,
    describe_stage_review_enabled,
    run_describe_stage_review,
)
from core.describe_json_qa import (
    SCHEMA_VERSION as JSON_QA_SCHEMA_VERSION,
    build_bailian_response_format_json_schema,
    coerce_answers_payload_by_stage_qa,
    coerce_answers_payload_defaults,
    parse_answers_json,
    render_answers_to_markdown,
    validate_answers_payload,
)
from core.per_llm_stages import (
    PLANNER_RECURSION_LIMIT,
    run_llm_planning_agent,
)

langchain.debug = True

from core.hf_env import apply_hf_hub_env_defaults

apply_hf_hub_env_defaults()

OUTPUT_DIR = "./output"

# 注入 Execute 任务：约束范围与顺序（对齐 Cursor scoped agent）
STAGE_EXECUTION_CONTRACT = """
---
## 【执行契约 — 已锁定计划】
上文「阶段执行计划」与「须按序完成的执行步骤」为本阶段**唯一**范围与顺序；各章专题要求以该章 prompt 为准。
1. **严格按 execution_steps** 调用工具并组织正文，勿跳步、勿倒序。
2. 未在步骤中点名的话题勿擅自铺成长篇；优先满足 must_cover 与 evidence_targets。
3. **证据**：关键结论须带可核验源码路径（反引号 `路径:行号`）；证据不足写「未发现」或「待核实」，禁止无路径断言。
4. **README / 文档（分层使用）**：
   - **允许且鼓励**：用 `list_repo_structure`、`read_code_segment` 阅读根目录 `README.md`、`docs/`、`*.md` 等，提取**如何构建**、**如何运行**（含 QEMU/板级命令）、依赖、目录说明、作者**声称**的评测/CI/课程环境。
   - **禁止**：仅凭 README、`doc/*.md` 或仓外文档断言「某内核机制已在代码中完整实现」；实现与否必须以源码+LSP/`grep` 为准。
   - **对照义务**：当 README 声称与代码相关时，须写「README 声称 vs 代码实际」并各附路径；无源码证据则标「待核实」或「未发现」。
5. **评测/交付适配信号（与本阶段相关时）**：若本章涉及构建、启动、块设备、网络、自测主控或产物交付，应优先用 `grep_in_repo` 等在仓库内检索计划所列启发式信号（如 `Makefile`/`all` 目标、固定内核产物名、自测标记字符串、`scripts/`、CI 配置、多 virtio 等）；**仅在有命中或 README 明确描述且能用源码佐证时**在正文写一小段「评测适配/提交契约」归纳并列出证据文件；否则可写「未发现仓库内评测专用 glue」或省略。
6. **桩与不可验证数据**：若符号仅为 `return 0`/`ENOSYS`/`todo!`/空体，须标为桩并引用片段；禁止编造具体 commit hash、未在当次 Git 工具输出中出现的短 SHA，或无法 `git`/工具复核的统计口径。
7. **文风**：原理点到即止，笔墨在实现；忌用大段「典型教学内核/通用教程」代正文。若提及外部框架，一两句带过，随即回到本仓代码。
8. **粒度**：机制落到字段、枚举分支、`#[cfg(feature)]`、锁与错误路径、分发表与关键调用边；**复杂分支先列 if/平台宏再下结论**，描述须与 `read_code_segment` 一致。
9. **平台与板级（本仓库）**：凡涉及启动、构建、驱动、根文件系统与块设备的步骤，须用 `list_repo_structure` / `grep_in_repo` 覆盖 `platform/`、`boards/`、`board*`、`dts/`、`dtb/`、`*.dts*`、`*.dtsi`、`board.toml`、`*defconfig*`、链接脚本；检索 `compatible`、`model`、`CONFIG_.*BOARD`、`CONFIG_.*MACH` 等与 **QEMU `-machine` / 多板型** 的绑定；命中则写差异并附 `路径:行号`，无命中则写「未发现」。
---
"""

import openpyxl
def normalize_url(url: str) -> str:
    """标准化仓库 URL，剔除 .git 后缀、结尾斜杠并统一小写"""
    if not url:
        return ""
    url = str(url).strip().lower()
    # 先剔除结尾斜杠
    while url.endswith("/"):
        url = url[:-1]
    # 再剔除 .git 后缀
    if url.endswith(".git"):
        url = url[:-4]
    # 再次剔除可能出现的结尾斜杠（例如 .git/ 情况）
    while url.endswith("/"):
        url = url[:-1]
    return url


def get_author_info(repo_url: str) -> dict:
    """从 collected-data.xlsx 提取作者信息"""
    try:
        xlsx_path = "collected-data.xlsx"
        if not os.path.exists(xlsx_path):
            return {}
            
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sheet = wb.active
        
        # 寻找列索引
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
                
        repo_col = headers.get("仓库地址")
        if not repo_col:
            return {}
            
        for row in range(2, sheet.max_row + 1):
            cell_url = sheet.cell(row=row, column=repo_col).value
            if cell_url and normalize_url(cell_url) == normalize_url(repo_url):
                result = {}
                mapping = {
                    "year": "年份",
                    "competition": "赛事",
                    "sub_competition": "子赛事",
                    "school": "学校",
                    "team": "队伍名称"
                }
                for key, col_name in mapping.items():
                    col_idx = headers.get(col_name)
                    if col_idx:
                        val = sheet.cell(row=row, column=col_idx).value
                        result[key] = str(val).strip() if val is not None else ""
                return result
    except Exception as e:
        print(f"  ⚠️  读取 collected-data.xlsx 失败: {e}")
    return {}


def _slug(s: str) -> str:
    keep = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_"):
            keep.append(ch)
        elif ch.isspace():
            keep.append("_")
    out = "".join(keep).strip("_")
    return out[:60] if out else "section"


def _build_base_context(repo_url: str, output_dir: str) -> str:
    """构建基础上下文信息。
    
    Args:
        repo_url: 仓库 URL
        output_dir: 输出目录
    """
    repo_name = repo_name_from_url(repo_url)
    # 与 tools.git_ops.get_repo_local_path 保持一致：./repos/<repo_name>
    repo_path = os.path.normpath(os.path.join("./repos", repo_name))
    charts_dir = os.path.normpath(os.path.join(output_dir, "charts"))
    sections_dir = os.path.normpath(os.path.join(output_dir, "sections"))
    
    # 后续阶段直接使用 repo_path
    repo_hint = f"**仓库已就绪**: 直接使用 repo_path = \"{repo_path}\"。"

    return f"""你是一个操作系统项目的技术分析 Agent。请严格基于仓库中的代码与文档输出结论，避免空泛；写作时使用《Operating Systems Internals and Design Principles（Stallings）》的专业术语与抽象层次框架。

基础信息：
- 仓库 URL: {repo_url}
- 本地路径 repo_path: {repo_path}
- 输出目录 output_dir: {output_dir}
- 图表目录 charts_dir: {charts_dir}
- 分段输出目录 sections_dir: {sections_dir}

{repo_hint}

全局术语与抽象层框架（必须遵守）：
0. **抽象层面坐标系**：任何机制描述都要先判定属于哪一层，避免“跨层混写/偷换概念”。
   - **硬件/平台层**：CPU 特权级与指令集 (ISA)、寄存器/特权指令、MMU/TLB、缓存/内存层次、设备控制器、总线、DMA、中断控制器。
   - **内核机制层**：中断/异常 (interrupt/exception)、陷入 (trap)、系统调用 (system call)、上下文切换 (context switch/dispatcher)、同步与互斥 (mutual exclusion)、调度 (scheduling)、内存管理与虚拟内存 (virtual memory/paging/segmentation)、I/O 子系统与驱动 (device drivers)、文件系统与命名 (file system/VFS)、网络 (networking)、多处理器 (SMP/AMP)。
   - **OS 服务/接口层**：进程/线程抽象 (process/thread)、地址空间 (address space)、文件描述符与 I/O 抽象、IPC、系统调用 ABI/接口语义、错误码约定。
   - **保护与安全层**：用户态/内核态隔离 (user mode/kernel mode)、保护域 (protection domain)、访问控制 (access control)、特权检查路径、可信计算基 (TCB) 边界与攻击面。
   - **部署与可运行性层**：引导链 (bootstrapping)、固件/引导加载器交接、构建配置/条件编译、平台适配矩阵、可观测性 (logging/debug/trace) 与故障模式。
1. **术语优先级**：优先使用教材常用术语并在首次出现时可括注英文：
   - 进程 (process)、线程 (thread)、进程/线程控制块 (PCB/TCB)、调度器 (scheduler)、分派器 (dispatcher)、时间片 (time slice)、抢占 (preemption)
   - 中断 (interrupt)、异常 (exception)、陷入 (trap)、陷入帧/陷入上下文 (trap frame)、系统调用 (system call)
   - 虚拟地址/物理地址 (VA/PA)、页 (page)、页帧 (frame)、页表 (page table)、TLB、缺页 (page fault)
   - 设备驱动 (device driver)、设备控制器 (device controller)、MMIO/PIO、DMA、缓冲/缓存 (buffer/cache)
   - 文件系统 (file system)、目录/目录项 (directory/dentry)、索引节点 (inode)、挂载 (mount)、VFS
   - 保护/安全 (protection/security)、访问控制 (ACL/capability 如适用)、特权级 (privilege level)
2. **覆盖性要求（跨章节一致）**：每章开头的 2-4 句必须明确：
   - 本章关注的**抽象层**（从上面的坐标系中选）；
   - 本章的**对象**（机制/接口/数据结构/路径）与**证据类型**（源码/LSP/RAG/构建配置）；
   - 若缺失证据：必须写“未发现/未实现/仅桩接口”并说明搜索范围与关键词。

全局要求（严格遵守）：
1. **反向证据原则**：如果未找到某功能的实现代码，你必须明确说明“未发现”或“未实现”。**严禁**仅仅因为它是“操作系统”就假设它实现了某些标准功能（如分页、多户）。
2. **证据引用**：描述任何关键结论（如“支持分页”）时，必须附带文件路径或代码片段引用（如 `mm/page.rs: map_page()`）。
3. **语义发现与拓扑追踪**：
   - 🔍 **第一优先级：语义切入**。在分析任何子系统前，**必须首先调用 `rag_search_code` 进行语义搜索**（例如：“寻找页表映射的实现”）。这能帮你穿透复杂的目录结构，直接定位到最相关的代码块。
   - 🌳 **第二优先级：拓扑展开**。通过 RAG 获得核心符号后，立即使用 `lsp_get_call_graph` 展开多层递归调用树，利用 `lsp_get_definition` 等 LSP 工具构建精确的 AST 画布。
   - 🛠️ **第三优先级：降级与查漏**。仅当 RAG 和 LSP 均返回为空或需要确切宏定义时，才触发 `grep_in_repo`。读取代码片段必须使用 `read_code_segment` 且仅限于关键逻辑。
4. **LSP 退避策略（必读）**：当 LSP 不可用或超时时，`lsp_get_definition` / `lsp_get_references` 会自动退避（Tree-sitter → 语言感知正则 → 通用 grep → ASM）。退避结果会附带 `[Fallback Metadata]`，含 `confidence=high|medium|low`。当 `confidence=low` 或结果含 `[Generic Fallback]` / `[ASM Fallback]` 时，在报告中必须标注「以上为静态分析结果，精度有限」，与 Call Graph 的 Grep 降级标注保持一致。
5. **多模块搜索**：不要局限于单一目录。使用 `find_os_core_modules` 寻找分散的实现（例如驱动可能在 `drivers/` 也可能在 `modules/`）。
6. **区分规划与实现**：README 中提到的功能可能是“画饼”，必须通过代码验证。未能验证的特性即使出现在文档中，也必须标注为“文档提及但未见代码”。
7. **桩代码检测（Strict Stub Detection）**：
   - 遇到函数体为空、返回 `unimplemented!()`、`todo!()`、`ENOSYS` 或仅有一行 `Ok(0)` 的情况，**必须**标注为 **“桩函数”** 或 **“未实现”**。
   - **严禁**将桩函数描述为“已实现功能”。如果一个系统调用仅返回 0 而无实际逻辑（如 `sys_getuid` 始终返回 0），必须指出“仅有接口无实现”。
8. **文件路径验证（Anti-Hallucination）**：
   - 在引用文件路径前，**必须**确保该文件在 `list_repo_structure` 或 `find_by_name` 的结果中真实存在。
   - **严禁**捏造不存在的文件路径（如声称 `riscv/boot.rs` 存在但实际不在）。如果不确定，不要写路径。
9. **头文件 vs 实现**：
   - 不要把头文件（`.h`）或 Trait 定义（`.rs` 中的 `trait`）作为功能已实现的证据。必须找到对应的 C 文件（`.c`）或 Rust `impl` 代码块。
10. **产出**：最终回答为完整 Markdown；执行计划未列出的工具步骤勿自行加戏。文风与粒度细则见每轮附带的执行契约。

输出使用 Markdown，面向懂 OS 的读者：原理点到即止，实现写透（路径与符号）。

**Markdown 格式规范**（严格遵守）：
1. **标题层级**：
   - 本章节已有一级标题（由系统自动添加），请从二级标题（##）开始
   - 禁止重复一级标题（#）
   - 标题层级递进：## → ### → ####，不要跳级
   - 禁止使用错误格式如 `### ##`

2. **正文输出**：
   - 直接输出分析内容，不要包含"现在我开始..."、"让我总结..."等思考过程
   - 不要输出分析前的准备文字

3. **代码块**：
   - 使用三个反引号包围，并标注语言（如 ```rust、```c）
   - 代码片段控制在50行以内，过长时使用注释省略

4. **列表**：
   - 有序列表使用 `1.`、`2.`，无序列表使用 `-` 或 `*`
   - 保持同级列表标记一致

5. **文件路径引用**：
   - 使用反引号包围：`arceos/modules/axtask/src/task.rs`
   - 在分析时总是给出完整相对路径

6. **代码元素引用**：
   - 结构体/函数/变量使用反引号：`TaskInner`、`spawn_task()`
   - 配置项使用反引号：`sched_fifo`

7. **图表引用**：
   - 使用标准 Markdown 格式：`![描述](charts/filename.png)`

8. **专业性**：
   - 使用准确的OS术语（如"页表"非"分页表"、"互斥锁"非"互斥"）
   - 中英文混排时注意空格（`进程 (Process)`）
"""


def _build_json_qa_prompt(stage_id: str, stage_title: str) -> tuple[str, list[str]]:
    """从 core/describe_stage_qa/<stage_id>.json 组装本阶段 prompt，并返回 expected_question_ids。"""
    stage_qa = load_stage_qa(stage_id)
    questions = stage_qa.get("questions", []) if isinstance(stage_qa, dict) else []
    expected_ids = list_question_ids(stage_qa)
    stage_constraints_md = stage_qa.get("stage_constraints_md", "") if isinstance(stage_qa, dict) else ""
    if not isinstance(stage_constraints_md, str):
        stage_constraints_md = ""

    # 题单（评委友好）+ 必须回答（供 planner must_cover 抽取）
    # + 阶段强相关约束（从 legacy prompt 迁移，避免“只剩题目没剩规范”）
    # + 输出契约（强约束 JSON）
    lines: list[str] = []
    lines.append("## 题单（评委用，小题作答）")
    lines.append("")
    for q in questions if isinstance(questions, list) else []:
        if not isinstance(q, dict):
            continue
        qid = str(q.get("question_id", "")).strip()
        qtype = str(q.get("question_type", "")).strip()
        stem = str(q.get("stem", "")).strip()
        if not (qid and qtype and stem):
            continue
        lines.append(f"- **{qid}** ({qtype})：{stem}")
        choices = q.get("choices")
        if isinstance(choices, list) and choices:
            for i, c in enumerate(choices):
                label = chr(ord("A") + i)
                lines.append(f"  - {label}. {str(c).strip()}")
    lines.append("")

    lines.append("## 必须回答")
    lines.append("")
    for q in questions if isinstance(questions, list) else []:
        if not isinstance(q, dict):
            continue
        qid = str(q.get("question_id", "")).strip()
        stem = str(q.get("stem", "")).strip()
        if qid and stem:
            lines.append(f"- [{qid}] {stem}")
    lines.append("")

    # 等价实现原则 + 停止条件（来自外部 QA 最佳实践）
    lines.append("## 分析原则（必须遵守）")
    lines.append("")
    lines.append(
        "**等价实现原则**：题目中出现的算法名、函数名、结构体名、宏名、系统调用名，"
        "都只是候选关键词或示例，**不是仓库里必须存在的精确标识符**。"
        "如果仓库采用不同命名，只要语义等价，就应按\u201c等价实现 / 等价接口 / 等价结构\u201d分析，"
        "**不要持续追逐不存在的名字**，不要因为特定函数名缺失就直接报 `not_found`。"
    )
    lines.append("")
    lines.append(
        "**停止条件**：一旦你已经能够稳定回答当前问题并闭合一条"
        "\u201c\u5165\u53e3 \u2192 \u6838\u5fc3\u64cd\u4f5c\uff08\u9875\u8868/\u5206\u914d/\u9501/\u8c03\u5ea6\u7b49\uff09\u201d"
        "的主链路，就**立即停止扩大搜索范围**，"
        "不要横向罗列大量相邻能力。已验证负向结论时写\u201c未发现实现\u201d，"
        "**不要在无直接证据时强行写\u201c未实现\u201d**。"
    )
    lines.append("")

    # 阶段强相关约束：一次性迁移到题库 JSON 中，运行时不再读取 legacy_stage_prompts。
    if stage_constraints_md.strip():
        lines.append("## 阶段强相关约束")
        lines.append("")
        lines.append(stage_constraints_md.strip())
        lines.append("")

    lines.append("## 输出契约（严格）")
    lines.append("")
    lines.append("你在本阶段**最终输出**必须是**唯一一个 JSON 对象**（允许使用 ```json 围栏包裹），不得输出任何额外解释文字。")
    lines.append("JSON 顶层字段必须包含：`schema_version`、`stage_id`、`stage_title`、`terminology_profile`、`answers`。")
    lines.append("其中 `answers` 是数组，每个元素必须包含：`question_id`、`question_type`、`stem`、`value`、`evidence`。")
    lines.append("`tri_state_impl` 的 `value` 只允许：`implemented` / `stub` / `not_found`。")
    lines.append("### 选择题约束（必须严格遵守，否则视为无效答案）")
    lines.append("")
    lines.append("**single_choice（单选）**：")
    lines.append("- `value` 必须是该题 `choices[]` 中的**某一项的完整原文**，要求**逐字匹配**（完全相等）。")
    lines.append("- **禁止**输出 `A/B/C/D` 或 `A.`/`B.` 前缀；禁止输出 `1/2/3` 编号；禁止改写/同义改写/缩写；禁止附加解释。")
    lines.append("- 如果你倾向于输出“字母+选项内容”（如 `A. xxx`），请改为只输出 `xxx`（与 choices 原文完全一致）。")
    lines.append("- 若证据不足以确定选哪一项，必须选择 `choices[]` 中语义最接近且包含 **“未发现/待核实/不支持/未实现”** 的那一项（若存在）。")
    lines.append("")
    lines.append("**multi_choice（多选）**：")
    lines.append("- `value` 必须是数组（JSON array）。数组中**每个元素**都必须是该题 `choices[]` 中的**完整原文**，要求逐字匹配。")
    lines.append("- **禁止**输出 `A/C`、`A,C`、`[\"A\",\"C\"]` 这类字母代号；禁止输出带 `A.` 前缀的文本；禁止改写选项。")
    lines.append("- 若题目允许“未发现/无统计”等选项，且你未找到任何证据支持其他项，优先把该“未发现/无统计”选项作为数组唯一元素。")
    lines.append(
        "每条证据 `evidence[]` 必须包含：`path`（仓库相对路径）、`symbol_kind`、`symbol_name`、"
        "`excerpt`（**否定/未找到结论时见上节须为非空**，其它题型仍建议填写摘录）。"
    )
    lines.append("")
    lines.append(
        "你必须使用本仓库代码证据作答。若未发现实现，按三态要求输出 `not_found`。"
        "**不得编造**不存在的文件路径或符号；但**鼓励**在 `evidence` 中给出**可复核的查找过程**（见下节），避免 `evidence` 为空导致评审无法判断你是否真的搜过。"
    )
    lines.append("")
    lines.append("### `not_found` / 否定结论时的证据（强烈建议非空）")
    lines.append("")
    lines.append(
        "当结论为 `not_found` 或选择题/简答表达「未发现」时，`evidence` **优先使用 1～3 条**记录检索过程（仍须真实、可对照）："
    )
    lines.append("- `path`：你实际打开或作为检索范围的**真实**仓库相对路径（可为具体文件，或代表性目录如 `kernel/syscall`；禁止虚构路径）。")
    lines.append("- `symbol_kind`：可用 `search`、`scan`、`doc` 等自说明类别（字符串即可）。")
    lines.append("- `symbol_name`：简短标签，如 `grep_pattern`、`syscall_table_scan`、`dir_list`。")
    lines.append(
        "- `excerpt`：**必填（此场景下）**，写清检索手段、关键字/模式、范围、结果（例如「于 sysnum.h 与 syscall.c 检索 socket/bind/listen 标识符，0 命中」或「已阅读 README 相关段落，未声称网络协议栈」）。"
    )
    lines.append("")
    lines.append("### Mermaid / 多行文本与 JSON 合法性（极其重要）")
    lines.append("")
    lines.append(
        "最终输出必须是 **`json.loads` 一次即可解析** 的合法 JSON。字符串值内的 **物理换行必须写成 `\\n`**，"
        "字符串内的 **双引号必须写成 `\\\"`**，反斜杠写成 `\\\\`。"
    )
    lines.append(
        "**严禁**在 `value` 的 JSON 字符串中直接嵌入「含真实换行」的 ```mermaid … ``` 围栏块——这是已引发 **Unterminated string / JSON 解析失败** 的高频原因。"
    )
    lines.append("若题干要求 Mermaid，任选其一（由易到难）：")
    lines.append("1. 将整图放在**单行字符串**内，仅用 `\\n` 表示图中换行（整条 `value` 对外仍是合法 JSON）；或")
    lines.append("2. 使用**无三反引号**的缩略 `graph TD` 文本（节点写 `Func[path:line]`），避免反引号围栏；或")
    lines.append("3. 用**有序步骤 + 路径:行号**列表代替图，并在 `notes`（若输出该字段）说明「为避免破坏 JSON 未嵌完整 Mermaid」。")
    lines.append("**禁止**输出「看起来像 JSON、实际不能解析」的内容；宁可简化图，不可输出非法 JSON。")
    lines.append("")
    lines.append("你必须按题单顺序输出 answers，且 question_id 必须与题单完全一致（不多不少）。")
    lines.append("")
    lines.append("JSON 示例（缩略，仅示意字段；不要复用其中路径/符号）：")
    lines.append("")
    lines.append("```json")
    lines.append("{")
    lines.append(f"  \"schema_version\": \"{JSON_QA_SCHEMA_VERSION}\",")
    lines.append(f"  \"stage_id\": \"{stage_id}\",")
    lines.append(f"  \"stage_title\": \"{stage_title}\",")
    lines.append("  \"terminology_profile\": \"stallings_en_zh\",")
    lines.append("  \"answers\": [")
    lines.append("    {")
    lines.append("      \"question_id\": \"QXX_001\",")
    lines.append("      \"question_type\": \"tri_state_impl\",")
    lines.append("      \"stem\": \"...\",")
    lines.append("      \"value\": \"not_found\",")
    lines.append("      \"evidence\": [")
    lines.append("        {")
    lines.append("          \"path\": \"kernel/syscall/sysnum.h\",")
    lines.append("          \"symbol_kind\": \"search\",")
    lines.append("          \"symbol_name\": \"grep_keyword_scan\",")
    lines.append(
        "          \"excerpt\": \"在 sysnum.h 与 syscall.c 检索关键字 xxx，0 命中；并列出已打开文件范围\""
    )
    lines.append("        }")
    lines.append("      ]")
    lines.append("    }")
    lines.append("  ]")
    lines.append("}")
    lines.append("```")
    lines.append("")

    return "\n".join(lines).strip() + "\n", expected_ids


STAGES = [

    {
        "id": "02_boot_trap",
        "title": "启动/架构与 Trap/系统调用",
        "prompt": "",  # JSON-QA: core/describe_stage_qa/02_boot_trap.json
    },
    {
        "id": "03_mem_mgmt",
        "title": "内存管理（物理/虚拟/分配器）",
        "prompt": "",
    },
    {
        "id": "04_process_smp",
        "title": "进程/线程/调度与多核",
        "prompt": "",
    },
    {
        "id": "05_fs_drivers",
        "title": "文件系统与设备 I/O",
        "prompt": "",
    },
    {
        "id": "06_sync_ipc",
        "title": "同步互斥与进程间通信",
        "prompt": "",
    },
    {
        "id": "07_security",
        "title": "安全机制与权限模型",
        "prompt": "",
    },
    {
        "id": "08_network",
        "title": "网络子系统与协议栈",
        "prompt": "",
    },
    {
        "id": "09_debug_error",
        "title": "调试机制与错误处理",
        "prompt": "",
    },
    {
        "id": "10_history",
        "title": "开发历史与里程碑",
        "prompt": """目标：撰写「开发历史与里程碑」一章，以 **Git 与仓库内可复核证据** 为主，做技术向演进归纳（避免无依据的提交流水账）。

**建议覆盖（有证据才写，无则明确写「未发现」）**
- 若存在 `.git`：用本阶段已启用的 Git 工具概括主分支/标签/作者贡献方向；**禁止**编造未出现在工具输出中的 commit hash 或统计口径。
- **模块演进**：可结合前文各章已出现的路径，用 `trace_file_evolution` 等追踪关键文件/目录的引入或大改，只写工具输出可支撑的判断。
- **文档里程碑**：README、CHANGELOG、CI 配置等中与版本或阶段目标相关的陈述，区分「文档声称」与「代码可验证结论」。
- **缺口**：若历史信息显示某能力被移除、长期标注实验/待办，如实记录。

**输出形式**：Markdown 本章正文（小节自洽）；**不要**以题库 JSON 对象作为终稿。""",
    },
    {
        "id": "01_overview",
        "title": "项目概览与技术栈",
        "needs_previous_sections": True,
        "prompt": """目标：在「项目概览」一章中，**以各前置章节（02–10）已生成的报告为主要依据**，系统整理**每个功能模块所采用的技术、算法、关键数据结构、外部依赖与工具使用**；并补充仓库级技术栈与构建信息。要求**逐模块覆盖、不得遗漏**（与下文「[前面阶段的分析内容]」中的章节一一对应）。

        **定位**：区分仓库对外名称与 README 中可能提及的上游框架名；若声称基于某框架，**一两句**交代关系即可，正文不展开框架教程。
        本阶段**最后执行**，上下文已含 **02–10** 全文；归纳以各章报告为主，仅在核对入口、架构列表、语言版本时**轻量**调工具。

        **模块—章节对应关系（归纳时必须逐条覆盖；若某章在上下文中缺失或过短，须单独说明「该章无可用内容」并写「待结合仓库核实」）**：
        - 02 启动/架构与 Trap/系统调用 → 启动链、模式切换、MMU/FPU、Trap 向量、系统调用分发、TrapFrame、用户指针校验等
        - 03 内存管理 → 物理/虚拟内存、页表、分配器、CoW/Lazy/Swap 等高级特性
        - 04 进程/线程/调度与多核 → 任务模型、调度算法、上下文切换、信号/Futex、多核启动、IPI、每 CPU 数据等
        - 05 文件系统与设备 I/O → VFS、具体 FS、fd/mmap、驱动框架、块/网设备等
        - 06 同步与 IPC → 锁、管道、IPC、Futex 等
        - 07 安全机制 → 权限、Capability、隔离等
        - 08 网络协议栈 → socket、协议栈实现形态等
        - 09 调试与错误处理 → panic、日志、回溯等
        - 10 演进与历史 → 模块演进、重大提交方向（技术层面归纳即可）

        请按顺序完成（repo_path 为仓库根）：
        1) **通读前置报告**：完整阅读上下文中的各章 Markdown，为下面「各模块技术全景」打草稿；摘录每章出现的**具体技术名词、算法名、crate 名、关键路径/符号**（须能在该章正文中找到依据，勿臆造）。
        2) analyze_tech_stack(repo_path)：核对/补充语言、构建、依赖与框架判断。
        3) list_repo_structure(repo_path, max_depth=5)：核对关键目录与入口线索。
        4) **架构支持**：列出本仓库实际支持的架构（可与前置报告交叉验证）。
        5) **内核入口**：用 LSP/grep 等**轻量**手段确认主入口符号与文件路径即可。
        6) **评测与交付信号（轻量）**：`read_code_segment` 阅读根目录 `README.md` 前几屏（若存在则扫 `docs/*.md` 标题区）提取构建/运行/QEMU 命令与作者**声称**的评测或 CI 环境；`grep_in_repo` 检索 `kernel-rv|kernel-la|disk.img|testcode|OS COMP|autograde|gitlab-ci|\\.github/workflows` 等中性模式。仅当有命中或 README 明确描述且能与前置章节或源码锚点交叉时写入下一节；否则在该节写「未发现仓库内评测专用适配信号」。

        输出格式要求（**严格按此顺序**，不得调换；`## 各模块技术全景` 不得省略）：

        - ## 快速总览
          **报告第一个内容块**。含：
          **一句话定位**（≤60字）："[OS名称] 基于 [框架/自研] 的 [架构] [内核类型]，主要语言 [x]，[1 个最突出技术点]。"
          **子系统完成度矩阵**（固定 10 行，不得删行；「关键实现」优先引用前置章节已给出的结论/路径；与 02–09 章对应时注明章节号）：
          | 子系统 | 完成度 | 关键实现 |
          |--------|--------|---------|
          | 启动与 Trap/系统调用（第 02 章） | ✅完整 / 🔸部分 / ❌缺失 | 一句话 |
          | 内存管理（第 03 章） | ... | ... |
          | 进程/调度与多核（第 04 章） | ... | ... |
          | 中断与系统调用（与第 02 章同源时可互引） | ... | ... |
          | 文件系统与设备 I/O（第 05 章） | ... | ... |
          | 同步与IPC（第 06 章） | ... | ... |
          | 多核支持（与第 04 章同源时可互引） | ... | ... |
          | 网络协议栈（第 08 章） | ... | ... |
          | 安全机制（第 07 章） | ... | ... |
          | 调试与错误处理（第 09 章） | ... | ... |

        - ## 评测与交付适配
          **固定小节、紧接在「快速总览」矩阵之后**。不臆测任何一届具体赛题或保密测例；仅综合 **02–09 技术前置章** 与 **本步对 README/`grep` 的轻量结果**，用短段落归纳下列四类（无证据则写「未发现」或「不适用」）：
          - **Delivery**：`Makefile`/`build.rs`/CI 中是否出现固定产物名（如 `kernel-rv`、`kernel-la`、`disk.img`）、`make all` 等；列路径，不写仓外断言。
          - **Harness**：是否存在自测主控、固定输出标记、扫盘跑脚本、关机等**输出契约**相关代码或 README 描述；须能指向前置章节或源码 `路径:行号`，否则标待核实。
          - **PlatformProfile**：README/QEMU 命令与代码中 **QEMU `virt` vs 物理板**、多 virtio、SMP 等是否一致；与 **02/05/04** 等章结论对照。
          - **SubsystemDepth**：若 README 声称可跑 libc/LTP/压测等，用 **02/05/08** 等章的桩态结论概括**风险缺口**（仍禁止数值打分）。
          禁止编造未在当次工具输出中出现的 commit hash。

        - ## 各模块技术全景（基于 02–10 章报告提取）
          **本章核心**：按上表「模块—章节对应」顺序，**每一模块一个小节**；**模块标题**须带章节号 `02`–`10` 与简短主题（与下表括号内主题一致即可）。

          **标题层级（强制，避免出现 `#### ### 02` 这类错误）**：
          - 先写本节总标题 `## 各模块技术全景（基于 02–10 章报告提取）`（若本章已用 `## 第 n 章` 包裹，则总标题用 `### …`，以**低于**章标题一级为准）。
          - **模块小节标题**须比「各模块技术全景」**再低一级**：例如总标题为 `###` 时，模块用 `#### 02 启动/架构与 Trap/系统调用`；总标题为 `##` 时，模块用 `### 02 …`。
          - **四个子节标题**须比模块标题**再低一级**：例如模块为 `####` 时，子节用 `##### 技术清单`、`##### 关键实现、证据与细粒度锚点`、`##### 依赖与工具`、`##### 与相邻模块的衔接`；模块为 `###` 时，子节用 `#### …`。
          - **禁止**在标题行 `#` 之后的文字里再写字面量 `###` / `##`（错误示例：`#### ### 02 …`）；**正确示例**：`#### 02 启动/架构与 Trap/系统调用`。

          **小节结构 Schema（强制，禁止省略或合并）**：每个模块小节下 **必须** 依次出现下列 **四个子节标题**（标题文字 **逐字** 如下；`#` 个数按上条层级规则），**不得**把「关键实现」「依赖」「衔接」并入「技术清单」段落或混写在固定条目行里。

          - `技术清单`（作为标题行的标题文本，勿含额外 `#`）
          - `关键实现、证据与细粒度锚点`
          - `依赖与工具`
          - `与相邻模块的衔接`

          **各段最低要求**：
          - **技术清单**（**固定维度，见下表**）：条列该技术模块中明确采用的技术/机制；**每条必须使用「固定条目名：归纳内容」**，其中 **「固定条目名」须与下表完全一致、顺序一致**（不得增删条目、不得改字、不得换序）；冒号后写本仓库/前置章可支撑的归纳（一句即可）。若某条目在本仓库不适用或无实现，冒号后写「不适用」并**简短**说明依据（如「前置第 0x 章未检出相关代码」），**不得**省略该条目行。**禁止**在技术清单条目中堆砌 `路径:行号`（证据只属于下一节）。
          - **关键实现、证据与细粒度锚点**：摘自前置第 0x 章报告，须能在该章正文找到依据，勿臆造。形式为项目符号列表；**至少 4 条**，且 **至少 3 条** 含 `源码路径:行号` 或 `` `路径` `` + `` `符号/结构体/字段` `` 级锚点（第 **10** 章可用 commit hash、作者、文档路径代替「源码行号」，但仍须 ≥4 条可复核证据）。可与技术清单一一对应，但**必须单独成段**。
          - **依赖与工具**：**至少 1 条** bullet；无第三方依赖时写「无外部 crate/库依赖」或「仅标准工具链与构建系统」等；有 LwIP/OpenSBI 等须点名版本或目录。**禁止**整段留空或仅写「无」一字。
          - **与相邻模块的衔接**：**至少 2 句**，且须显式提到 **至少两个** 相邻章节编号（如「与第 03 章页表切换…」「为第 06 章管道提供…」）；第 02 章可衔接 03/04/06 等，第 10 章可衔接 03–09 中至少两处。**禁止**用「见上文」代替。

          若某模块在前置报告中内容极多，仅在 **「关键实现、证据与细粒度锚点」** 下再分子标题，**不得**省略四个 `####` 中的任一段。

          **技术清单「固定条目名」表（冒号前必须逐字使用；每模块恰好下列行数）**：
          条目命名取 **教材/手册级抽象**（机制与模型），**不**绑定某一具体 OS 实现细节；冒号后归纳本仓库。若某条完全不适用，冒号后写「不适用」并简述依据，**不得**删行。

          - **### 02（启动/架构与 Trap/系统调用）**（7 行）：
            `启动链与引导交接：` / `特权级与执行模式（硬件隔离模型）：` / `MMU 与内核地址空间初建：` / `同步异常与用户态陷阱入口（含 syscall 路径）：` / `异步设备中断与中断控制器抽象：` / `时钟源与定时中断（tick/计账/抢占触发）：` / `用户内存访问与系统调用参数安全（copyin/out 等）：`
          - **### 03（内存管理）**（6 行）：
            `物理内存组织与页帧分配器：` / `页表、地址空间与虚实地址转换：` / `缺页与页面错误处理（含按需分页/惰性路径）：` / `进程虚拟地址空间布局与映射接口：` / `高级策略（CoW/Lazy/换页/mmap 等）：` / `页缓存或与 FS 块缓存的边界（归入本章或与第 05 章交叉说明）：`
          - **### 04（进程/调度与多核）**（6 行）：
            `进程或线程抽象与调度实体（PCB/TCB）：` / `调度策略与就绪队列结构：` / `抢占模型与时间片/优先级（可协作则注明）：` / `上下文切换与内核栈/寄存器约定：` / `生命周期（创建/执行/阻塞/退出/wait 与僵尸）：` / `多核、每 CPU 状态与 IPI/迁移（若适用）：`
          - **### 05（文件系统与设备 I/O）**（6 行）：
            `VFS 与 inode/file 等对象模型：` / `路径解析与挂载/命名空间：` / `具体文件系统实现形态：` / `文件描述符与打开文件表：` / `块缓存、写回与磁盘 I/O 路径：` / `字符设备与块设备驱动框架（含 virtio 等）：`
          - **### 06（同步与 IPC）**（6 行）：
            `自旋锁与中断上下文临界区规则：` / `可睡眠互斥与锁序/死锁约束（若述及）：` / `等待队列、睡眠与唤醒：` / `管道等字节流 IPC：` / `信号与异步通知：` / `共享内存或 futex 等（若本仓库有）：`
          - **### 07（安全机制）**（5 行）：
            `硬件隔离与特权域模型：` / `访问控制模型（DAC/MAC/Capability 等，无则写不适用）：` / `用户指针验证与内核/用户空间数据拷贝边界：` / `可执行空间保护与权限位策略（W^X 等）：` / `其他沙箱或策略（seccomp/namespace/cgroup 等，无则写不适用）：`
          - **### 08（网络协议栈）**（5 行）：
            `套接字抽象与用户态 API：` / `协议栈分层与数据面实现形态：` / `网卡驱动与收发包/DMA 路径：` / `与协议栈缓冲与 sk_buff 类抽象（若适用）：` / `与文件层或块设备的衔接（若适用）：`
          - **### 09（调试与错误处理）**（5 行）：
            `Panic/oops 与致命错误停机路径：` / `日志级别与可观测输出：` / `栈回溯与符号化/调试钩子：` / `断言与运行时检查：` / `系统调用级追踪或 strace 类能力：`
          - **### 10（演进与历史）**（4 行）：
            `活跃时间范围与提交规模：` / `核心贡献者与模块分工：` / `重大重构或技术里程碑：` / `文档与工程化沉淀：`
          上表中的 `/` 仅用于本说明排版；**正式输出中须写成 7～6～6～6～6～5～5～5～4 条独立列表项**（每模块内一行一条「固定条目名：…」），**不得**用斜杠把多条挤在同一行。

        - ## 技术栈与构建（编程语言版本、框架、依赖、支持的架构完整列表）
        - ## 目录结构导读（关键目录与源码入口）
        - ## 总结评价（完成度评估）
          200-300 字：结合**各模块技术全景**与完成度矩阵，概括闭环情况；**禁止数值打分**（不要出现 x/10 等）。

        **收尾**：**「各模块技术全景」**须脱离 02–10 原文仍可读懂各模块技术选型；**且**每个 `02`–`10` 模块小节均须含 **四个**子节（技术清单 / 关键实现… / 依赖与工具 / 与相邻模块的衔接），**缺一不可**，不得仅输出「技术清单」；**标题 `#` 级数须符合上文「标题层级」规则**，**禁止**出现 `#### ### 02` 或在标题文本中嵌入 `###`。
        """,
    },
]

# _format_tool_call_summary 和 _format_tool_result_summary 已移至 core.utils 模块


def _strip_llm_preamble(text: str) -> str:
    """剥掉 LLM 输出中第一个 Markdown 标题（# 开头）之前的过渡性思考文字。

    LLM 有时会在正式报告内容之前输出一段口语化的过渡文字，例如：
      "现在我已经收集了足够的信息来撰写...让我整理分析结果..."
    这类文字不属于报告内容，需要在写入文件前过滤掉。

    策略：找到第一个以 '#' 开头的行，从该行开始作为有效内容。
    如果全文均无 '#' 标题行，则保留原文（避免误删真实内容）。
    """
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if line.strip().startswith('#'):
            stripped = "\n".join(lines[i:]).strip()
            if i > 0:
                dropped = "\n".join(lines[:i]).strip()
                # 只在确实有内容被剥掉时打印提示
                if dropped:
                    preview = dropped[:120].replace('\n', ' ')
                    print(f"  ✂️  已剥除 LLM 前缀（{len(dropped)} 字符）: {preview}...")
            return stripped
    # 没有找到 '#' 标题，保留原文
    return text


def print_step(step_num: int, node_name: str, state: dict, stage_step_num: int = 0, max_steps: int = 1500, stage_limit: int = 600) -> int:
    """打印每一步的执行信息（简洁的 agent 风格）
    
    Args:
        step_num: 全局步骤号
        node_name: 节点名称
        state: 状态字典
        stage_step_num: 阶段内步骤号
        max_steps: 全局最大步数估计（仅显示用）
        stage_limit: 阶段递归步数限制
        
    Returns:
        int: 本次步骤消耗的 token 数量
    """
    token_count = 0
    messages = state.get("messages", [])
    if not messages:
        return 0
    
    # 只显示新增的消息（避免重复）
    if step_num == 1:
        msg_to_show = messages
    else:
        msg_to_show = [messages[-1]] if messages else []
    
    for msg in msg_to_show:
        if isinstance(msg, AIMessage):
            content = msg.content or ""
            tool_calls = getattr(msg, "tool_calls", None) or []
            
            if step_num > 0:
                print(f"\n【Total Step {step_num}/{max_steps}】(Stage step: {stage_step_num}/{stage_limit})", end=" ")

            if tool_calls:
                print("🔧 Tool Calls:")
                for tc in tool_calls:
                    if isinstance(tc, dict):
                        tool_name = tc.get("name", "unknown")
                        tool_args = tc.get("args", {})
                    else:
                        tool_name = getattr(tc, "name", "unknown")
                        tool_args = getattr(tc, "args", {})
                    
                    summary = format_tool_call_summary(tool_name, tool_args)
                    print(f"   {tool_name}({summary})")
            
            # 如果有思考内容且没有工具调用，显示思考（这通常是最终输出）
            elif content.strip():
                # 显示简短预览
                preview = content.strip()[:200]
                if len(content) > 200:
                    preview += "..."
                print(f"Agent: {preview}")
            
            # 打印 Token Usage
            metadata = getattr(msg, "response_metadata", {})
            usage = metadata.get("token_usage", {})
            if usage:
                total_this_call = usage.get("total_tokens", 0)
                if total_this_call > 0:
                    token_count += total_this_call
        
        elif isinstance(msg, ToolMessage):
            tool_name = getattr(msg, "name", "unknown")
            content = msg.content or ""
            summary = format_tool_result_summary(tool_name, content)
            print(f"   ✅ {tool_name}: {summary}")
            
    return token_count


def _save_json(path: str, payload: dict):
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        import json
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  ⚠️  无法写入 JSON 文件 {path}: {e}")


def _invoke_describe_stage_review(
    repo_output_dir: str,
    stage_id: str,
    title: str,
    expected_question_ids: list,
    payload_before_stage_qa: dict,
    skip_in_report: bool,
) -> None:
    """JSON-QA 校验成功且非 01/10 时，送审题单 + 题库覆写前的答案 JSON。"""
    if skip_in_report or not describe_stage_review_enabled():
        return
    if not describe_stage_review_applies(stage_id, expected_question_ids=expected_question_ids):
        return
    sidecar_dir = os.path.join(repo_output_dir, "_per_stage")
    os.makedirs(sidecar_dir, exist_ok=True)
    question_sheet = build_stage_qa_question_sheet(stage_id, title)
    model_json = json.dumps(payload_before_stage_qa, ensure_ascii=False, indent=2)
    try:
        parsed_review, review_raw_llm, review_err = run_describe_stage_review(
            stage_id=stage_id,
            stage_title=title,
            question_sheet=question_sheet,
            model_json_before_stage_qa_coerce=model_json,
            expected_question_ids=list(expected_question_ids),
        )
        review_out = os.path.join(sidecar_dir, f"{stage_id}_review.json")
        review_err_out = os.path.join(sidecar_dir, f"{stage_id}_review_error.json")
        if parsed_review is not None:
            parsed_review["_meta"] = {
                "review_model": os.environ.get("DESCRIBE_REVIEW_MODEL") or os.getenv("MODEL_NAME", ""),
            }
            _save_json(review_out, parsed_review)
            conf = parsed_review.get("confidence")
            print(f"\n   📋 Describe Review：confidence={conf} → {review_out}")
        else:
            _save_json(
                review_err_out,
                {
                    "error": review_err or "unknown",
                    "raw_model_output_excerpt": (review_raw_llm or "")[:8000],
                },
            )
            print(f"\n   ⚠️ Describe Review 解析失败（已写入 {review_err_out}）")
    except Exception as _re:
        _save_json(
            os.path.join(sidecar_dir, f"{stage_id}_review_error.json"),
            {"error": f"{type(_re).__name__}: {_re}"},
        )
        print(f"\n   ⚠️ Describe Review 调用失败: {_re}")


def _clear_stage_sidecar_artifacts(repo_output_dir: str, stage_id: str) -> None:
    """重新生成某阶段前删除 _per_stage 下该 stage_id 的旧 JSON，避免与本次 run 混用。"""
    sidecar_dir = os.path.join(repo_output_dir, "_per_stage")
    if not os.path.isdir(sidecar_dir):
        return
    prefix = f"{stage_id}_"
    removed: list[str] = []
    for name in os.listdir(sidecar_dir):
        if not name.startswith(prefix) or not name.endswith(".json"):
            continue
        path = os.path.join(sidecar_dir, name)
        try:
            os.remove(path)
            removed.append(name)
        except OSError:
            pass
    if removed:
        print(f"   🧹 已清理本阶段上次残留的侧车: {', '.join(sorted(removed))}")


def _summarize_section_text(text: str, max_chars: int = 500) -> str:
    text = _strip_llm_preamble((text or "").strip())
    text = re.sub(r"\n{3,}", "\n\n", text)
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3] + "..."


def _extract_path_mentions(text: str) -> list[str]:
    pattern = re.compile(r"`([^`\n]*/[^`\n]+\.(?:rs|c|h|cpp|go|zig|S|s|toml|ld|md|py)(?::\d+(?:-\d+)?)?)`")
    return [m.group(1) for m in pattern.finditer(text or "")]


def main():
    repo_url = os.environ.get("REPO_URL", "").strip()
    
    if not repo_url:
        print("❌ 错误：未设置 REPO_URL 环境变量")
        print("   请在 .env 文件中设置 REPO_URL，或通过命令行设置：")
        print("   export REPO_URL=\"https://github.com/example/os-project.git\"")
        sys.exit(1)

    repo_name = repo_name_from_url(repo_url)
    
    # 按 OS 名称创建独立的输出目录
    repo_output_dir = os.path.join(OUTPUT_DIR, repo_name)
    sections_dir = os.path.join(repo_output_dir, "sections")
    
    os.makedirs(repo_output_dir, exist_ok=True)
    os.makedirs(sections_dir, exist_ok=True)

    print("=" * 80)
    print("📋 多阶段任务：开始")
    print(f"Repo: {repo_name}")
    print(f"⏰ 开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)

    all_section_paths = []
    overall_step_count = 0
    total_tokens_used = 0  # 累计token使用量
    start_time = datetime.now()
    
    # 计算有效章节数（用于文件命名）
    chapter_counter = 0  # 实际章节计数器
    
    GLOBAL_ESTIMATED_STEPS = 1500  # 全局估计总步数（用于显示进度条分母）

    # 初始化错误追踪器
    error_tracker = ErrorTracker(output_dir=repo_output_dir)


    # 在此处增加直接克隆逻辑
    repo_local_path = os.path.normpath(os.path.join("./repos", repo_name))
    print("=" * 80)
    print(f"📦 阶段 0：仓库准备 (直接执行，无需 LLM)")
    print("=" * 80)
    if os.path.exists(repo_local_path) and os.path.isdir(repo_local_path) and os.listdir(repo_local_path):
        print(f"⏭️  仓库已存在，跳过克隆。 路径: {repo_local_path}")
    else:
        print(f"🚀 正在克隆仓库: {repo_url} ...")
        from tools.git_ops import clone_repository
        result = clone_repository.invoke({"repo_url": repo_url})
        print(result)

    # 新跑开始前：删除仓库根带 OS-Agent 签名的遗留临时文件（compile_flags / 虚拟 Cargo 等）
    try:
        from tools.lsp_ops import cleanup_os_agent_repo_ephemeral

        cleanup_os_agent_repo_ephemeral(repo_local_path)
    except Exception as _e:
        print(f"⚠️  清理历史 LSP 临时文件跳过: {_e}")
    
    # --- 增加：RAG 预索引 (加速后续分析阶段) ---
    print(f"\n🚀 阶段 0.5：RAG 预索引 (代码向量化)...")
    try:
        from core.code_rag import CodeRAGEngine
        # 注意：这里传入的 repo_local_path 是 ./repos/xv6-k210
        rag_engine = CodeRAGEngine(project_name=repo_name)
        # build_index 会自动检测并跳过已存在的索引
        rag_engine.build_index(repo_local_path, force=False)
        print(f"✅ RAG 预索引完成，后续语义搜索将秒开。")
    except Exception as e:
        print(f"⚠️ RAG 预索引跳过 (将在首次调用时重试): {e}")
    # ------------------------------------------

    repo_profile = build_repo_profile(repo_url=repo_url, repo_path=repo_local_path)
    _save_json(os.path.join(repo_output_dir, "repo_profile.json"), repo_profile)
    global_memory = {
        "section_summaries": {},
        "mentioned_paths": [],
        "external_background": {},
    }

    for idx, stage in enumerate(STAGES, 1):
        stage_id = stage["id"]
        title = stage["title"]
        prompt = stage["prompt"]
        expected_question_ids: list[str] = []
        # JSON-QA：仅当 STAGES 未写内联 prompt（prompt 为空）时，才从 describe_stage_qa 组装题单并可能覆盖 prompt。
        # 内联阶段（如 01_overview、10_history）不读题库、不要求 JSON 答案后处理。
        if not (prompt or "").strip():
            try:
                qa_prompt, expected_question_ids = _build_json_qa_prompt(stage_id=stage_id, stage_title=title)
                if expected_question_ids:
                    prompt = qa_prompt
            except Exception:
                expected_question_ids = []
        stage_state = StageState(
            stage_id=stage_id,
            stage_title=title,
            stage_type="describe",
            stage_prompt=prompt,
        )


        # 检查是否跳过此阶段
        skip_in_report = stage.get("skip_in_report", False)
        
        # 只有非skip阶段才计入章节号
        if not skip_in_report:
            # 优先从 stage_id 提取章节号 (例如 "02_boot_trap" -> "02")
            chapter_prefix = stage_id.split('_')[0]
            if not chapter_prefix.isdigit():
                chapter_counter += 1
                chapter_prefix = f"{chapter_counter:02d}"
            else:
                chapter_counter = int(chapter_prefix)
                
            section_name = f"{chapter_prefix}_{_slug(title)}.md"
        else:
            # skip阶段使用idx作为前缀（避免冲突，但不会保存）
            section_name = f"00_{_slug(title)}.md"
        
        section_path = os.path.join(sections_dir, section_name)

        # 简单的断点续传：如果文件已存在且内容看起来正常（>200字节），则跳过
        if os.path.exists(section_path):
            if os.path.getsize(section_path) > 200:
                print("=" * 80)
                print(f"⏭️  阶段 {idx}/{len(STAGES)}：{title} (已存在，跳过)")
                print(f"   文件: {section_path}")
                print("=" * 80)
                if not skip_in_report:
                    all_section_paths.append(section_path)
                    try:
                        with open(section_path, "r", encoding="utf-8", errors="ignore") as existing_f:
                            existing_text = existing_f.read().strip()
                        global_memory["section_summaries"][stage_id] = _summarize_section_text(existing_text)
                        global_memory["mentioned_paths"].extend(_extract_path_mentions(existing_text))
                        global_memory["mentioned_paths"] = list(dict.fromkeys(global_memory["mentioned_paths"]))[-100:]
                    except Exception:
                        pass
                continue
            else:
                print(f"♻️  检测到残留的失败文件 (Size: {os.path.getsize(section_path)} bytes)，将删除并重试: {section_name}")
                try:
                    os.remove(section_path)
                except OSError:
                    pass

        _clear_stage_sidecar_artifacts(repo_output_dir, stage_id)

        # 每个阶段构建 base_ctx（先打出阶段标题，再 Plan，避免日志上像「先 Plan 再阶段一」）
        base_ctx = _build_base_context(repo_url=repo_url, output_dir=repo_output_dir)
        print("\n" + "=" * 80)
        print(f"🧩 阶段 {idx}/{len(STAGES)}：{title}")
        print("=" * 80)
        print("\n" + "━" * 26 + " ① Plan · 探索并锁定本阶段计划 " + "━" * 26+'\n')
        stage_state.plan = plan_stage(stage_state, repo_profile=repo_profile, global_memory=global_memory)
        try:
            planner_agent = build_planner_agent(stage_id=stage_id)
            plan_stream_step = 0

            def _on_plan_stream(node_name: str, st: dict) -> None:
                nonlocal overall_step_count, plan_stream_step
                overall_step_count += 1
                plan_stream_step += 1
                print_step(
                    overall_step_count,
                    node_name,
                    st,
                    plan_stream_step,
                    GLOBAL_ESTIMATED_STEPS,
                    PLANNER_RECURSION_LIMIT,
                )

            plan_overlay, plan_notes = run_llm_planning_agent(
                planner_agent,
                stage_state,
                repo_profile,
                repo_local_path,
                global_memory=global_memory,
                on_stream_step=_on_plan_stream,
            )
            stage_state.metadata["llm_plan_overlay_applied"] = bool(plan_overlay)
            if plan_overlay:
                stage_state.plan = apply_llm_plan_overlay(stage_state.plan, plan_overlay)
            if plan_notes:
                stage_state.metadata["llm_plan_notes"] = plan_notes
            if plan_overlay or plan_notes:
                print(f"   🧩 LLM 规划已合并（overlay_keys={list(plan_overlay.keys()) if plan_overlay else []}）")
        except Exception as e:
            stage_state.metadata["llm_plan_overlay_applied"] = False
            logging.warning("LLM 规划失败，沿用启发式计划: %s", e)
        stage_state.plan = ensure_execution_steps(stage_state.plan)
        stage_state.dynamic_context = build_dynamic_context(stage_state, repo_profile=repo_profile, global_memory=global_memory)
        plan_context = render_plan_context(stage_state)
        
        # 如果是最后整合阶段，需要读取前面所有 section 的内容
        previous_sections_content = ""
        if stage.get("needs_previous_sections", False) and all_section_paths:
            print(f"\n📚 读取前面 {len(all_section_paths)} 个阶段的分析内容...")
            sections_texts = []
            
            # 限制每个 section 的最大字符数，避免 "lost in the middle"；01 概览需做全模块技术归纳，单章上限略放宽
            max_chars_per_section = 28000 if stage_id == "01_overview" else 17000
            total_chars = 0
            
            for sp in all_section_paths:
                try:
                    with open(sp, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read().strip()
                        if content:
                            orig_len = len(content)
                            # 如果内容太长，截取前面部分 + 提示
                            if orig_len > max_chars_per_section:
                                truncated = content[:max_chars_per_section]
                                # 尝试在段落边界截断
                                last_para = truncated.rfind("\n\n")
                                if last_para > max_chars_per_section * 0.7:
                                    truncated = truncated[:last_para]
                                content = truncated + f"\n\n... [此部分已截断，原文 {orig_len} 字符]"
                            
                            sections_texts.append(f"--- 来自: {os.path.basename(sp)} ---\n{content}")
                            total_chars += len(content)
                except Exception as e:
                    print(f"  ⚠️  无法读取 {sp}: {e}")
            
            if sections_texts:
                previous_sections_content = "\n\n" + "=" * 60 + "\n[前面阶段的分析内容摘要]\n" + "=" * 60 + "\n\n"
                previous_sections_content += "\n\n".join(sections_texts)
                previous_sections_content += "\n\n" + "=" * 60 + "\n[以上是前面阶段的分析内容，请基于这些内容整合生成最终报告]\n" + "=" * 60 + "\n"
                print(f"  ✅ 已加载 {len(sections_texts)} 个 section，共 {total_chars} 字符（每个最多 {max_chars_per_section} 字符）")
        
        # 从已解析的 chapter_counter 直接获取章节号
        chapter_num = chapter_counter if not skip_in_report else None
        
        # Execute：单条 Human（计划 + 写作任务合并），随后进入 ReAct
        if chapter_num:
            chapter_hint = f"""
**本阶段是最终报告的第 {chapter_num} 章：{title}**

请按照以下格式输出本章内容（这将直接作为最终报告的一部分）：

## 第 {chapter_num} 章：{title}

（你的分析内容）

"""
            human_plan_block = (
                base_ctx
                + "\n"
                + chapter_hint
                + "\n"
                + plan_context
                + "\n"
                + STAGE_EXECUTION_CONTRACT
            )
        else:
            human_plan_block = (
                base_ctx
                + "\n"
                + f"## 阶段 {idx}/{len(STAGES)}：{title}\n\n"
                + plan_context
                + "\n"
                + STAGE_EXECUTION_CONTRACT
            )
        human_write_block = (
            "## 本阶段须撰写的分析内容（在遵守上文计划与执行契约的前提下完成）\n\n"
            + prompt
            + previous_sections_content
        )
        combined_execute_prompt = human_plan_block.rstrip() + "\n\n" + human_write_block
        inputs = {
            "messages": [
                SystemMessage(content=DESCRIBE_SYSTEM_PROMPT_JSON),
                HumanMessage(content=combined_execute_prompt),
            ]
        }

        print("\n\n   ✅ ① Plan 已锁定 → 执行步骤条数:", len(stage_state.plan.execution_steps) if stage_state.plan else 0)
        print("\n" + "━" * 26 + " ② Execute · 按锁定计划执行 " + "━" * 26+'\n')
        print(f"   模型: {os.getenv('MODEL_NAME')} ")
        sys.stdout.flush()  # 强制刷新输出缓冲区

        # 若本阶段为题库 JSON-QA：
        # 不要在 ReAct 执行阶段启用 response_format/json_schema。
        # 原因：启用结构化输出会让 LangChain 进入“自动解析工具调用”的模式，
        # 要求所有 function tools 都是 strict；而本项目工具（如 get_repo_local_path）并非 strict，
        # 会直接报：`<tool>` is not strict. Only strict function tools can be auto-parsed。
        #
        # JSON-QA 的结构化约束交给后处理（extract/validate）完成。
        agent = build_executor_agent(stage_id=stage_id, model_kwargs=None)
        
        final_state = None
        stage_step_count = 0  # 阶段内步骤计数
        recursion_limit = 600
        
        stage_tokens = 0  # 阶段内 token 计数
        retry_count = 0
        last_exception = None
        
        # 智能重试循环
        while retry_count <= RetryConfig.MAX_RETRIES:
            try:
                if retry_count > 0:
                    # 分类错误并决定是否重试
                    error_type = classify_error(last_exception)
                    
                    if not RetryConfig.RETRYABLE_ERRORS.get(error_type, False):
                        logging.warning(
                            f"[{title}] 错误类型 {error_type.value} 不适合重试，跳过"
                        )
                        break
                    
                    # 计算退避时间
                    backoff = calculate_backoff(retry_count - 1)
                    print(f"   🔄 正在重试 ({retry_count}/{RetryConfig.MAX_RETRIES})...")
                    print(f"   ⏱️  等待 {backoff} 秒后重试（{error_type.value}）...")
                    time.sleep(backoff)
                    
                    logging.info(
                        f"[{title}] 第 {retry_count} 次重试 "
                        f"(错误类型: {error_type.value}, 退避: {backoff}s)"
                    )
                
                # 执行 agent
                for event in agent.stream(inputs, config={"recursion_limit": recursion_limit}):
                    overall_step_count += 1
                    stage_step_count += 1
                    for node_name, state in event.items():
                        step_tokens = print_step(overall_step_count, node_name, state, stage_step_count, GLOBAL_ESTIMATED_STEPS, recursion_limit)
                        stage_tokens += step_tokens
                        final_state = state
                
                # 成功执行完成，跳出重试循环
                logging.info(f"[{title}] Agent 执行成功（步骤数: {stage_step_count}）")
                break
                
            except KeyboardInterrupt:
                print("\\n\\n⚠️  用户中断执行")
                error_tracker.save_error_report(filename="describe_error_report.json")
                sys.exit(1)
            except Exception as e:
                last_exception = e
                error_type = classify_error(e)
                
                # 记录错误到追踪器
                error_tracker.record_error(
                    section_name=title,
                    error_type=error_type,
                    exception=e,
                    retry_count=retry_count,
                    context={
                        "stage_id": stage_id,
                        "step_count": stage_step_count,
                        "model": os.getenv('MODEL_NAME'),
                        "recursion_limit": recursion_limit,
                    }
                )
                
                print(f"\n   ❌ {error_type.value}: {type(e).__name__}: {e}")
                retry_count += 1
                
                if retry_count > RetryConfig.MAX_RETRIES:
                    error_msg = f"超过最大重试次数 {RetryConfig.MAX_RETRIES}"
                    logging.critical(
                        f"[{title}] {error_msg} - 最后错误: {type(e).__name__}: {e}"
                    )
                    import traceback
                    traceback.print_exc()
                    # 不退出，继续下一阶段

        stage_text = ""
        is_complete = False
        all_ai_content = []  # 收集所有 AI 回复内容
        execution_messages = []
        
        if final_state and final_state.get("messages"):
            messages = final_state["messages"]
            execution_messages = messages
            if messages:
                # 收集所有 AIMessage 的内容（降低阈值到 20 字符）
                for m in messages:
                    if isinstance(m, AIMessage):
                        content = (m.content or "").strip()
                        if content and len(content) > 20:
                            all_ai_content.append(content)
                
                # 优先查找：最后一条包含实际内容且没有工具调用的 AIMessage
                for m in reversed(messages):
                    if isinstance(m, AIMessage):
                        content = (m.content or "").strip()
                        tool_calls = getattr(m, "tool_calls", None) or []
                        
                        # 理想情况：有内容且没有工具调用（这是最终总结）
                        if content and not tool_calls and len(content) > 200:
                            print(f"\n✅ 找到有效 AI 回复（无工具调用，长度: {len(content)} 字符）")
                            stage_text = content
                            is_complete = True
                            break
                
                # 如果没找到理想的最终回复，尝试合并所有 AI 内容
                if not is_complete and all_ai_content:
                    print(f"\n⚠️  未找到独立最终回复，合并 {len(all_ai_content)} 条 AI 回复内容")
                    # 取最长的那条
                    stage_text = max(all_ai_content, key=len)
                    # 如果最长的不够长，合并所有内容
                    if len(stage_text) < 500 and len(all_ai_content) > 1:
                        stage_text = "\n\n---\n\n".join(all_ai_content)
                
                # 如果仍然没有内容，且该阶段需要生成报告，则发送追问消息
                if (not stage_text or len(stage_text) < 100) and not skip_in_report:
                    print(f"\n🔄 内容不足，发送追问消息请求生成报告...")
                    followup_msg = HumanMessage(content=f"""你已经收集了关于"{title}"的信息。

现在请根据你收集到的信息，立即输出完整的 Markdown 格式分析报告。

不要再调用任何工具，直接写出报告内容。报告必须包含：
1. 设计概览
2. 核心数据结构
3. 关键实现分析
4. 代码片段引用

请现在输出报告：""")
                    
                    try:
                        # 必须清理掉末尾带有未解析 tool_calls 的 AIMessage，否则大模型会抛出验证错误
                        # 并导致部分模型（如 Qwen）重置上下文。
                        safe_messages = messages.copy()
                        while safe_messages and isinstance(safe_messages[-1], AIMessage) and getattr(safe_messages[-1], "tool_calls", None):
                            safe_messages.pop()
                            
                        # 继续对话，追加 followup 消息
                        followup_inputs = {"messages": safe_messages + [followup_msg]}
                        followup_final_state = None
                        for event in agent.stream(followup_inputs, config={"recursion_limit": 10}):
                            overall_step_count += 1
                            stage_step_count += 1
                            for node_name, state in event.items():
                                step_tokens = print_step(overall_step_count, node_name, state, stage_step_count, GLOBAL_ESTIMATED_STEPS, recursion_limit)
                                stage_tokens += step_tokens
                                followup_final_state = state
                                # 提取追问后的回复
                                if state.get("messages"):
                                    for m in reversed(state["messages"]):
                                        if isinstance(m, AIMessage):
                                            content = (m.content or "").strip()
                                            tool_calls = getattr(m, "tool_calls", None) or []
                                            if content and not tool_calls and len(content) > 100:
                                                print(f"\n✅ 追问后获得有效回复（长度: {len(content)} 字符）")
                                                stage_text = content
                                                is_complete = True
                                                break
                                    if is_complete:
                                        break
                            if is_complete:
                                break
                        # 始终用追问流式最后一帧的完整 messages，保留先前所有 ToolMessage 供证据抽取
                        if followup_final_state and followup_final_state.get("messages"):
                            execution_messages = followup_final_state["messages"]
                    except Exception as e:
                        print(f"\n⚠️  追问失败: {e}")
                
                if not stage_text:
                    print(f"\n❌ 未找到任何有效 AI 回复内容。")
                    last_msg = messages[-1]
                    stage_text = f"> ⚠️ **生成警告**: 该章节未能完整生成. Last msg type: {type(last_msg).__name__}"

        if not stage_text.strip():
            stage_text = "> ⚠️ **生成警告**: Agent 未返回有效内容。"

        stage_text = _strip_llm_preamble(stage_text.strip())

        # --- JSON-QA 后处理：解析/校验/落盘答案 JSON，并渲染为 Markdown 供现有审阅器使用 ---
        if expected_question_ids:
            sidecar_dir = os.path.join(repo_output_dir, "_per_stage")
            os.makedirs(sidecar_dir, exist_ok=True)
            raw_path = os.path.join(sidecar_dir, f"{stage_id}_answers_raw.txt")
            json_path = os.path.join(sidecar_dir, f"{stage_id}_answers.json")
            try:
                with open(raw_path, "w", encoding="utf-8") as f:
                    f.write(stage_text.strip() + "\n")
            except Exception:
                pass

            try:
                stage_qa = load_stage_qa(stage_id)
                payload = parse_answers_json(stage_text)
                payload = coerce_answers_payload_defaults(payload)
                payload_before_stage_qa = copy.deepcopy(payload)
                payload = coerce_answers_payload_by_stage_qa(payload, stage_qa=stage_qa)
                issues = validate_answers_payload(
                    payload,
                    stage_id=stage_id,
                    stage_title=title,
                    expected_question_ids=expected_question_ids,
                )
                if issues:
                    _save_json(
                        os.path.join(sidecar_dir, f"{stage_id}_answers_validate_issues.json"),
                        {"issues": [{"path": x.path, "reason": x.reason} for x in issues]},
                    )
                else:
                    _save_json(json_path, payload)
                    stage_text = render_answers_to_markdown(payload).strip()
                    _invoke_describe_stage_review(
                        repo_output_dir,
                        stage_id,
                        title,
                        expected_question_ids,
                        payload_before_stage_qa,
                        skip_in_report,
                    )
            except Exception as e:
                # JSON 解析/校验失败：将“输出 JSON 前的全部对话 + 原输出”重发给模型，要求仅返回合法 JSON 进行修复重生成。
                # 注意：这一步必须禁止工具调用，否则会变成“继续探索”而不是“修复 JSON”。
                stage_qa = load_stage_qa(stage_id)

                repair_error = f"{type(e).__name__}: {e}"
                _save_json(
                    os.path.join(sidecar_dir, f"{stage_id}_answers_parse_error.json"),
                    {"error": repair_error},
                )

                try:
                    # 尽量保留上下文：使用执行阶段的完整 messages（包含工具证据），但避免把“未解析 tool_calls 的 AIMessage”带入。
                    safe_msgs = (execution_messages or []).copy()
                    while safe_msgs and isinstance(safe_msgs[-1], AIMessage) and getattr(safe_msgs[-1], "tool_calls", None):
                        safe_msgs.pop()

                    qa_prompt, _expected_ids = _build_json_qa_prompt(stage_id=stage_id, stage_title=title)
                    # 最多重试 3 次修复重生成
                    repair_llm = build_chat_model(temperature=0, max_retries=0)
                    last_repair_exc: Exception | None = None
                    for attempt in range(1, 4):
                        try:
                            repair_instructions = (
                                "上一次你生成的 JSON 无法被解析或无法通过校验。\n"
                                f"错误：{repair_error}\n"
                                f"修复重试次数：{attempt}/3\n\n"
                                "请你基于上文所有对话与工具证据，重新输出**唯一一个 JSON 对象**（允许 ```json 围栏），不要输出任何额外解释。\n"
                                "严格遵守题库的 question_id / question_type / stem。\n"
                                "- single_choice 的 value 必须是 choices 中的完整选项文本（禁止 A/B/C/D）。\n"
                                "- multi_choice 的 value 必须是数组，元素为 choices 中完整选项文本（禁止字母代号）。\n\n"
                                "这是题库与输出契约（必须按此修复）：\n\n"
                                + qa_prompt
                                + "\n\n"
                                "这是你上一次的原始输出（供你修复，可能包含多余文字/围栏/不完整 JSON）：\n\n"
                                "```text\n"
                                + (stage_text or "")
                                + "\n```\n"
                            )

                            repaired = repair_llm.invoke(safe_msgs + [HumanMessage(content=repair_instructions)])
                            repaired_text = (getattr(repaired, "content", "") or "").strip()

                            with open(raw_path, "a", encoding="utf-8") as f:
                                f.write(f"\n\n---\n# repair_attempt {attempt}/3\n")
                                f.write(repaired_text + "\n")

                            payload = parse_answers_json(repaired_text)
                            payload = coerce_answers_payload_defaults(payload)
                            payload_before_stage_qa = copy.deepcopy(payload)
                            payload = coerce_answers_payload_by_stage_qa(payload, stage_qa=stage_qa)
                            issues = validate_answers_payload(
                                payload,
                                stage_id=stage_id,
                                stage_title=title,
                                expected_question_ids=expected_question_ids,
                            )
                            if not issues:
                                _save_json(json_path, payload)
                                stage_text = render_answers_to_markdown(payload).strip()
                                _invoke_describe_stage_review(
                                    repo_output_dir,
                                    stage_id,
                                    title,
                                    expected_question_ids,
                                    payload_before_stage_qa,
                                    skip_in_report,
                                )
                                last_repair_exc = None
                                break
                            _save_json(
                                os.path.join(sidecar_dir, f"{stage_id}_answers_validate_issues.json"),
                                {"issues": [{"path": x.path, "reason": x.reason} for x in issues]},
                            )
                        except Exception as e_attempt:
                            last_repair_exc = e_attempt

                    if last_repair_exc is not None:
                        raise last_repair_exc
                except Exception as e2:
                    _save_json(
                        os.path.join(sidecar_dir, f"{stage_id}_answers_repair_error.json"),
                        {"error": f"{type(e2).__name__}: {e2}"},
                    )

        artifacts = extract_stage_artifacts(stage_text, execution_messages)
        stage_state.draft_markdown = artifacts["draft_markdown"] or stage_text
        stage_state.draft_document = artifacts["draft_document"]
        stage_state.evidence_index = artifacts["evidence_index"]
        stage_state.status = "executed"

        sidecar_dir = os.path.join(repo_output_dir, "_per_stage")
        _save_json(
            os.path.join(sidecar_dir, f"{stage_id}_plan.json"),
            stage_state.plan.to_dict() if stage_state.plan else {},
        )

        # 落盘章节正文
        if skip_in_report:
            print(f"\n⏭️  阶段 {idx} 标记为 skip_in_report，不写入报告")
        else:
            try:
                os.makedirs(os.path.dirname(section_path), exist_ok=True)
                with open(section_path, "w", encoding="utf-8") as f:
                    clean_text = _strip_llm_preamble(stage_state.draft_markdown.strip())
                    f.write(clean_text + "\n")
                all_section_paths.append(section_path)
                global_memory["section_summaries"][stage_id] = _summarize_section_text(clean_text)
                global_memory["mentioned_paths"].extend(_extract_path_mentions(clean_text))
                global_memory["mentioned_paths"] = list(dict.fromkeys(global_memory["mentioned_paths"]))[-100:]
                if stage_id == "01_overview":
                    background_items = [
                        item.excerpt[:300]
                        for item in stage_state.evidence_index
                        if item.source_type == "web_background" and item.excerpt
                    ]
                    if background_items:
                        global_memory["external_background"]["competition_background"] = background_items[:4]
                print(f"\n✅ 已保存阶段输出: {section_path}")
            except Exception as e:
                print(f"\n⚠️  无法写入阶段文件 {section_path}: {e}")

        # 统计本阶段的token使用（已在 print_step 中累加）
        # stage_tokens 已准备好
        
        total_tokens_used += stage_tokens
        if stage_tokens > 0:
            print(
                f"   阶段「{title}」token +{stage_tokens:,} "
                f"（本阶段 {stage_step_count} 次 LLM）→ 累计 {total_tokens_used:,}"
            )
            sys.stdout.flush()

    # 生成 Call Graph 概览块（缓存由 compile 配置/git/管线指纹自动失效，无需环境变量）
    callgraph_md = ""
    try:
        from tools.callgraph_overview import generate_callgraph_section
        callgraph_md, cg_llm_tokens = generate_callgraph_section(
            repo_path=repo_local_path,
            output_dir=repo_output_dir,
            top_k=30,
            use_embedding=True,
            lsp_refine=True,
            force_regenerate=False,
        )
        if cg_llm_tokens > 0:
            total_tokens_used += cg_llm_tokens
            print(
                f"   Call Graph（domain×layer）token +{cg_llm_tokens:,} "
                f"→ 累计 {total_tokens_used:,}"
            )
            sys.stdout.flush()
    except Exception as e:
        print(f"\n⚠️  Call Graph 生成失败: {e}")
        import traceback; traceback.print_exc()

    # 合并总报告 - 生成专业的、类似人类撰写的技术文档
    final_report_path = os.path.join(repo_output_dir, f"OS技术分析报告_{repo_name}.md")
    try:
        # 01_overview 排最前，其余按文件名排序
        overview_sections = [p for p in all_section_paths if os.path.basename(p).startswith("01_")]
        other_sections = sorted([p for p in all_section_paths if not os.path.basename(p).startswith("01_")])
        content_sections = overview_sections + other_sections

        with open(final_report_path, "w", encoding="utf-8") as out:
            # 标题和元数据
            out.write(f"# {repo_name} 操作系统技术分析报告\n\n")

            # 加载作者信息
            author_info = get_author_info(repo_url)
            if author_info:
                if author_info.get("year"):
                    out.write(f"> **年份**: {author_info['year']}\n\n")
                if author_info.get("competition"):
                    out.write(f"> **赛事**: {author_info['competition']}\n\n")
                if author_info.get("sub_competition"):
                    out.write(f"> **子赛事**: {author_info['sub_competition']}\n\n")
                if author_info.get("school"):
                    out.write(f"> **学校**: {author_info['school']}\n\n")
                if author_info.get("team"):
                    out.write(f"> **队伍名称**: {author_info['team']}\n\n")

            out.write(f"> **仓库地址**: {repo_url}\n\n")
            out.write(f"> **分析日期**: {datetime.now().strftime('%Y年%m月%d日')}\n\n")
            out.write(f"> **分析工具**: OS-Agent-D\n\n")
            out.write("---\n\n")

            # 目录
            out.write("## 目录\n\n")
            for i, p in enumerate(content_sections, 1):
                try:
                    filename = os.path.basename(p)
                    title = os.path.splitext(filename)[0]
                    if '_' in title:
                        title = title.split('_', 1)[1]
                    title = title.replace('_', ' ')
                    out.write(f"{i}. {title}\n")
                except Exception:
                    filename = os.path.splitext(os.path.basename(p))[0]
                    out.write(f"{i}. {filename}\n")

            out.write("\n---\n\n")

            # Call Graph 概览（放在目录之后、章节正文之前）
            if callgraph_md:
                out.write(callgraph_md)
                out.write("\n---\n\n")

            # 正文：依次输出各章节内容
            for i, p in enumerate(content_sections, 1):
                try:
                    filename = os.path.basename(p)
                    stem_fn = os.path.splitext(filename)[0]
                    chapter_heading = stem_fn.replace("_", " ")
                    if "_" in stem_fn:
                        prefix, rest = stem_fn.split("_", 1)
                        if prefix.isdigit():
                            chapter_heading = f"第{prefix}章 {rest.replace('_', ' ')}"

                    with open(p, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read().strip()

                        out.write(f"\n# {chapter_heading}\n\n")

                        if content.startswith("# "):
                            content = "##" + content[1:]

                        out.write(content + "\n\n")
                        out.write("---\n\n")
                except Exception as e:
                    print(f"  ⚠️  无法读取章节 {p}: {e}")

            # 页脚
            out.write(f"\n---\n\n")
            out.write(f"*本报告由 OS-Agent-D 自动生成*  \n")
            out.write(f"*生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*  \n")
            out.write(f"*分析耗时: {(datetime.now() - start_time).total_seconds()/60:.1f} 分钟*\n")

        print(f"\n📄 已生成最终报告: {final_report_path}")
        print(f"   报告包含 {len(content_sections)} 个主要章节")
    except Exception as e:
        print(f"\n⚠️  无法生成总报告: {e}")
        import traceback
        traceback.print_exc()

    # 保存错误报告（如果有错误）
    error_tracker.save_error_report(filename="describe_error_report.json")
    if error_tracker.errors:
        print(f"\n⚠️  运行过程中发生 {len(error_tracker.errors)} 个错误")
        print(error_tracker.generate_error_summary())

    end_time = datetime.now()
    elapsed = (end_time - start_time).total_seconds()

    print("=" * 80)
    print("✅ 多阶段任务完成！")
    print(f"   总步数: {overall_step_count}")
    print(f"   总Token使用: {total_tokens_used:,}")
    print(f"   耗时: {elapsed:.2f} 秒 ({elapsed/60:.2f} 分钟)")
    if error_tracker.errors:
        print(f"   错误数: {len(error_tracker.errors)} (详见 describe_error_report.json)")
    print(f"⏰ 结束时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)


if __name__ == "__main__":
    main()
