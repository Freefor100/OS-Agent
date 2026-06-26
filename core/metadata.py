"""Submission metadata from collected-data.xlsx — authoritative year/school/team.

Bridges xlsx entries to local repos/ via git remote URL matching.
Replaces the hardcoded FRAMEWORKS set in search.py and the fragile
regex-based _extract_year().

Usage:
    mm = MetadataManager()
    meta = mm.lookup_by_repo_name("T202410487992457-1800")
    # → {"year": 2024, "school": "清华大学", "team": "...", ...}
    mm.is_framework("xv6-riscv")  # → True (not in xlsx)
    mm.same_year("T202410487992457-1800")  # → list of same-year submissions
"""

from __future__ import annotations

import logging
import subprocess
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parents[1]


def _normalize_url(url: str) -> str:
    """Strip .git suffix, trailing slash, normalize to https://, lowercase."""
    url = url.strip().rstrip("/")
    if url.startswith("git@") and ":" in url:
        host, path = url[4:].split(":", 1)
        url = f"https://{host}/{path}"
    if url.endswith(".git"):
        url = url[:-4]
    for prefix in ("git://", "git+https://", "git+http://"):
        if url.startswith(prefix):
            url = "https://" + url[len(prefix):]
    if url.startswith("http://"):
        url = "https://" + url[7:]
    return url.lower()


class MetadataManager:
    """Load xlsx, scan repo git remotes, build URL↔repo_name index."""

    URL_COLUMNS = (
        ("初赛fork版", "初赛fork版访问情况"),
        ("源地址", "源地址访问情况"),
        ("决赛fork版", "决赛fork版访问情况"),
        # Backward compatibility for the old one-URL workbook.
        ("仓库地址", ""),
    )

    def __init__(self, xlsx_path: str = "", repos_dir: str = ""):
        self._xlsx_path = str(ROOT / xlsx_path) if xlsx_path else str(ROOT / "collected-data.xlsx")
        self._repos_dir = str(ROOT / repos_dir) if repos_dir else str(ROOT / "repos")

        # xlsx row by normalized URL. One workbook row can expose initial fork,
        # source, and final fork addresses, but consumers get row metadata only.
        self._by_url: dict[str, dict] = {}
        # repo_name → xlsx entry
        self._by_name: dict[str, dict] = {}
        # stable work key → xlsx base row
        self._works: dict[str, dict] = {}
        # repo_name → git remote URL (for repos not in xlsx)
        self._repo_urls: dict[str, str] = {}
        # set of repo names NOT in xlsx (frameworks/baselines)
        self._framework_names: set[str] = set()

        self._load_xlsx()
        self._scan_repos()

    # ── loading ──────────────────────────────────────────────────────

    def _load_xlsx(self) -> None:
        """Parse collected-data.xlsx → self._by_url."""
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for xlsx metadata. Install with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(self._xlsx_path, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {str(value).strip(): idx for idx, value in enumerate(header_row) if value}

        def get(row: tuple, name: str, default: str = ""):
            idx = headers.get(name)
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return default if value is None else value

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            year = get(row, "年份")
            competition = get(row, "赛事")
            sub_event = get(row, "子赛事")
            school = get(row, "学校")
            team = get(row, "队伍名称")
            if not any((year, competition, sub_event, school, team)):
                continue
            year_int = int(year) if str(year or "").strip().isdigit() else 0
            work_key = "|".join([
                str(year_int),
                str(sub_event or "").strip(),
                str(school or "").strip(),
                str(team or "").strip(),
            ])
            base = {
                "year": year_int,
                "competition": str(competition or ""),
                "sub_event": str(sub_event or ""),
                "school": str(school or ""),
                "team": str(team or ""),
                "work_key": work_key,
                "xlsx_row": row_idx,
                "address_note": str(get(row, "地址备注") or ""),
                "official_source": str(get(row, "官网来源") or ""),
                "competition_id": str(get(row, "比赛ID") or ""),
            }
            self._works[work_key] = base

            for url_col, access_col in self.URL_COLUMNS:
                url = get(row, url_col)
                if not url:
                    continue
                url = str(url).strip()
                if not url or url in {"空", "-", "无", "None", "none"}:
                    continue
                access = str(get(row, access_col) or "").strip() if access_col else ""
                norm = _normalize_url(url)
                self._by_url[norm] = {
                    **base,
                    "repo_url": url,
                    "access_status": access,
                }

    def _copy_entry(self, entry: dict, *, repo_name: str = "") -> dict:
        out = dict(entry)
        if repo_name:
            out["repo_name"] = repo_name
            out["local_repo"] = repo_name
        return out

    def _work_entries(self, work_key: str) -> list[dict]:
        return [entry for entry in self._by_url.values() if entry.get("work_key") == work_key]

    def _preferred_entry_for_work(self, work_key: str) -> Optional[dict]:
        entries = self._work_entries(work_key)
        return entries[0] if entries else None

    def _base_work_entry(self, work_key: str) -> dict:
        preferred = self._preferred_entry_for_work(work_key) or self._works.get(work_key, {})
        return {
            "year": preferred.get("year", 0),
            "competition": preferred.get("competition", ""),
            "sub_event": preferred.get("sub_event", ""),
            "school": preferred.get("school", ""),
            "team": preferred.get("team", ""),
            "work_key": work_key,
            "xlsx_row": preferred.get("xlsx_row", 0),
            "address_note": preferred.get("address_note", ""),
            "official_source": preferred.get("official_source", ""),
            "competition_id": preferred.get("competition_id", ""),
        }

    def _scan_repos(self) -> None:
        """Scan repos/*/ for git remotes, match to xlsx entries."""
        repos = Path(self._repos_dir)
        if not repos.is_dir():
            return

        for d in sorted(repos.iterdir()):
            if not d.is_dir() or d.name.startswith("."):
                continue
            url = self._get_remote_url(str(d))
            if not url:
                continue
            norm = _normalize_url(url)
            self._repo_urls[d.name] = url

            entry = self._by_url.get(norm)
            if entry:
                self._by_name[d.name] = self._copy_entry(entry, repo_name=d.name)
            else:
                self._framework_names.add(d.name)

    @staticmethod
    def _get_remote_url(repo_path: str) -> str:
        """git remote get-url origin for a repo directory."""
        try:
            r = subprocess.run(
                ["git", "-C", repo_path, "remote", "get-url", "origin"],
                capture_output=True, text=True, timeout=10,
            )
            return r.stdout.strip() if r.returncode == 0 else ""
        except Exception:
            return ""

    # ── public API ───────────────────────────────────────────────────

    def lookup_by_url(self, git_url: str) -> Optional[dict]:
        """Look up xlsx metadata by git remote URL. Returns None if not found."""
        entry = self._by_url.get(_normalize_url(git_url))
        return dict(entry) if entry else None

    def lookup_by_repo_name(self, repo_name: str) -> Optional[dict]:
        """Look up xlsx metadata by local repos/<name> directory name."""
        return self._by_name.get(repo_name)

    def lookup_by_repo_path(self, repo_path: str) -> Optional[dict]:
        """Resolve repo_path → git remote → xlsx lookup.

        repo_path can be:
          - "repos/T202410487992457-1800"
          - "T202410487992457-1800" (bare name)
          - an absolute path
        """
        # Try as directory name first
        name = Path(repo_path).name
        if name in self._by_name:
            return self._by_name[name]
        # Try git remote lookup
        url = self._get_remote_url(str(ROOT / repo_path) if "/" not in repo_path else repo_path)
        if url:
            entry = self._by_url.get(_normalize_url(url))
            return self._copy_entry(entry, repo_name=name) if entry else None
        return None

    def is_framework(self, repo_name: str) -> bool:
        """True if repo is NOT in xlsx → it's a framework, baseline, or teaching prototype."""
        return repo_name in self._framework_names

    def is_framework_by_path(self, repo_path: str) -> bool:
        """Check framework status by path (resolves git remote)."""
        name = Path(repo_path).name
        if name in self._framework_names:
            return True
        if name in self._by_name:
            return False
        return self.is_framework(name)

    def same_year_submissions(self, repo_name: str) -> list[dict]:
        """All non-framework submissions from the same year as repo_name."""
        entry = self._by_name.get(repo_name)
        if not entry:
            return []
        year = entry["year"]
        return [
            {"repo_name": name, **data}
            for name, data in self._by_name.items()
            if data["year"] == year and name != repo_name
        ]

    def all_submissions(self) -> list[dict]:
        """All xlsx-matched submissions with their repo names."""
        return [
            {"repo_name": name, **data}
            for name, data in sorted(self._by_name.items())
        ]

    def get_framework_names(self) -> set[str]:
        """Repo names NOT in xlsx. Replaces the hardcoded FRAMEWORKS set."""
        return self._framework_names.copy()

    def unmatched_xlsx_entries(self) -> list[dict]:
        """Workbook rows that have no matching local repo through any address."""
        return self.unmatched_works()

    def unmatched_url_entries(self) -> list[dict]:
        """Address cells in xlsx that have no matching local repo."""
        matched_urls = {_normalize_url(u) for u in self._repo_urls.values()}
        return [
            entry for url, entry in self._by_url.items()
            if url not in matched_urls
        ]

    def all_works(self) -> list[dict]:
        """All workbook rows, independent of how many URL variants each has."""
        return [
            self._base_work_entry(work_key)
            for work_key in sorted(self._works)
        ]

    def unmatched_works(self) -> list[dict]:
        """Workbook rows with none of their URL variants cloned locally."""
        matched_urls = {_normalize_url(u) for u in self._repo_urls.values()}
        rows: list[dict] = []
        for work_key in sorted(self._works):
            entries = self._work_entries(work_key)
            if entries and not any(_normalize_url(e["repo_url"]) in matched_urls for e in entries):
                rows.append(self._base_work_entry(work_key))
        return rows

    def get_all_branch_prefs(self, repo_name: str = "") -> dict:
        """Return relevant metadata for branch scoring context.
        Includes competition deadline knowledge for time-phase matching.
        """
        entry = self._by_name.get(repo_name, {})
        return {
            "year": entry.get("year", 0),
            "school": entry.get("school", ""),
            "team": entry.get("team", ""),
            "is_framework": repo_name in self._framework_names,
        }

    # ── stats ────────────────────────────────────────────────────────

    def stats(self) -> dict:
        """Summary statistics for diagnostics."""
        years: dict[int, int] = defaultdict(int)
        schools: dict[str, int] = defaultdict(int)
        for entry in self._by_name.values():
            years[entry["year"]] += 1
            schools[entry["school"]] += 1

        return {
            "xlsx_entries": len(self._works),
            "xlsx_address_entries": len(self._by_url),
            "matched_repos": len(self._by_name),
            "frameworks": len(self._framework_names),
            "unmatched_xlsx": len(self.unmatched_xlsx_entries()),
            "unmatched_address_entries": len(self.unmatched_url_entries()),
            "years": dict(sorted(years.items())),
            "school_count": len(schools),
            "repos_scanned": len(self._repo_urls),
        }


# ── CLI for diagnostics ─────────────────────────────────────────────

if __name__ == "__main__":
    mm = MetadataManager()
    s = mm.stats()
    print(f"xlsx entries:      {s['xlsx_entries']}")
    print(f"xlsx addresses:    {s['xlsx_address_entries']}")
    print(f"matched repos:     {s['matched_repos']}")
    print(f"frameworks:        {s['frameworks']}")
    print(f"unmatched xlsx:    {s['unmatched_xlsx']}")
    print(f"unmatched address: {s['unmatched_address_entries']}")
    print(f"years:             {s['years']}")
    print(f"schools:           {s['school_count']}")
    print()
    print("Frameworks:")
    for f in sorted(mm.get_framework_names()):
        print(f"  {f}")
    unmatched = mm.unmatched_xlsx_entries()
    if unmatched:
        print(f"\nUnmatched xlsx entries ({len(unmatched)}):")
        for e in unmatched[:5]:
            print(f"  {e['year']} {e['school']} {e['team']}")
